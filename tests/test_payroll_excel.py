import tempfile
import zipfile
import inspect
from io import BytesIO
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

from apps.dashboard import openai_usage, records, router as dashboard_router, timesheet_corrections, timesheet_parser, timesheet_uploads
from apps.dashboard.payroll_calculations import build_week_sheet_rows, derived_period_total_rows
from apps.dashboard.payroll_excel import analyze_payroll_workbook, build_payroll_output_workbook, build_tgn_template_output_workbook


try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill
except ImportError:  # pragma: no cover
    Workbook = None
    PatternFill = None


@unittest.skipIf(Workbook is None, "openpyxl is niet beschikbaar")
class PayrollExcelAnalysisTests(unittest.TestCase):
    def test_analyzes_period_five_reference_structure(self):
        workbook = Workbook()
        workbook.active.title = "Periode"
        workbook["Periode"]["A1"] = "Werknemer"
        workbook["Periode"]["B1"] = "Contracturen"
        workbook["Periode"]["C1"] = "Bruto uurloon"

        for sheet_name in ["WK17", "WK18", "WK19", "WK20"]:
            sheet = workbook.create_sheet(sheet_name)
            sheet["A1"] = "Werknemer"
            sheet["B1"] = "Uren gewerkt"
            sheet["C1"] = "Netto voorschot"
            sheet["D2"] = "=B2+C2"

        payslip = workbook.create_sheet("Loonstrook")
        payslip["A1"] = "werknemer"
        payslip["B1"] = "totale uren gewerkt"
        payslip["C1"] = "nog te ontvangen netto loon"
        workbook.create_sheet("Grondslag bouw & infra")
        workbook.create_sheet("SAVG")

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "TGN verloning 2026 Periode 5.xlsx"
            workbook.save(path)
            analysis = analyze_payroll_workbook(path)

        self.assertEqual([week["week_number"] for week in analysis["week_tabs"]], [17, 18, 19, 20])
        self.assertEqual(analysis["period_sheet"], "Periode")
        self.assertEqual(analysis["payslip_sheet"], "Loonstrook")
        self.assertIn("Grondslag bouw & infra", analysis["foundation_sheets"])
        self.assertGreaterEqual(analysis["formula_count"], 4)
        self.assertIn("worked_hours", analysis["mapped_fields"]["WK17"])


class PayrollCalculationTests(unittest.TestCase):
    def test_decimal_hours_keep_fraction_with_dot_or_comma(self):
        from apps.dashboard import payroll_calculations

        self.assertEqual(payroll_calculations._decimal("37.5"), payroll_calculations.Decimal("37.5"))
        self.assertEqual(payroll_calculations._decimal("37,5"), payroll_calculations.Decimal("37.5"))
        self.assertEqual(payroll_calculations._decimal("5.156,25"), payroll_calculations.Decimal("5156.25"))

    def test_derives_period_totals_from_existing_payroll_rows(self):
        rows = derived_period_total_rows(
            [
                {
                    "employee_name": "Thomas",
                    "worked_days": 4,
                    "total_hours": "32",
                    "gross_amount": "€ 800,00",
                }
            ]
        )

        self.assertEqual(rows[0]["employee_name"], "Thomas")
        self.assertEqual(rows[0]["total_worked_hours"], "32")
        self.assertEqual(rows[0]["total_period_amount"], "€ 800,00")
        self.assertEqual(rows[0]["status"], "concept")

    @unittest.skipIf(Workbook is None, "openpyxl is niet beschikbaar")
    def test_exports_workbook_tabs(self):
        period = {
            "workbook_tabs": [
                {
                    "label": "WK17",
                    "columns": [{"label": "Werknemer", "key": "employee_name"}],
                    "rows": [{"employee_name": "Thomas"}],
                },
                {
                    "label": "Periode",
                    "columns": [{"label": "Werknemer", "key": "employee_name"}],
                    "rows": [{"employee_name": "Thomas"}],
                },
                {
                    "label": "Loonstrook",
                    "columns": [{"label": "Werknemer", "key": "employee_name"}],
                    "rows": [{"employee_name": "Thomas"}],
                },
            ]
        }
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "output.xlsx"
            build_payroll_output_workbook(path, period)
            workbook = load_workbook(path)

        self.assertEqual(workbook.sheetnames, ["WK17", "Periode", "Loonstrook"])

    def test_week_rows_use_week_level_days_and_km_from_timesheet_parse(self):
        rows = build_week_sheet_rows(
            "WK21",
            [],
            [
                {
                    "employee_name": "Leo Doorn",
                    "week_hours": ["0", "0", "0", "40"],
                    "week_worked_days": ["0", "0", "0", "5"],
                    "week_total_km": ["0", "0", "0", "105"],
                    "week_timesheet_ids": [[], [], [], [123]],
                    "total_km": "105",
                }
            ],
            {"week_index": 4, "week_number": 21},
        )

        self.assertEqual(rows[0]["worked_days"], "5")
        self.assertEqual(rows[0]["worked_hours"], "40")
        self.assertEqual(rows[0]["total_km"], "105")
        self.assertEqual(rows[0]["commute_km"], "105")

    @unittest.skipIf(Workbook is None, "openpyxl is niet beschikbaar")
    def test_exports_tgn_template_layout_with_formulas(self):
        template = Workbook()
        template.active.title = "WK21"
        for sheet_name in ["WK22", "WK23", "WK24", "Periode", "Loonstrook", "Grondslag bouw & infra", "SAVG"]:
            template.create_sheet(sheet_name)
        template["Periode"]["B7"] = "Werknemer"
        template["Periode"]["B8"] = "Oud"
        template["Periode"]["G8"] = "40"
        template["Periode"]["N8"] = "=L8+M8"
        template["WK21"]["B7"] = "Werknemer"
        template["WK21"]["B2"] = "ALLEEN GEKLEURDE LETTERS EN CIJFERS INGEVEN!"
        template["WK21"]["B2"].fill = PatternFill("solid", fgColor="FFFF00")
        template["WK21"]["B8"] = "=Periode!B8"
        template["WK21"]["C8"] = "=Periode!G8"
        template["WK21"]["K8"] = "=(Periode!BB8*E8/40)"
        template["WK21"]["Q8"] = "=K8+O8+P8"
        template["WK21"]["R8"] = "oude opmerking"
        template["WK21"]["S8"] = "oude projectinfo"
        template["WK21"]["R8"].fill = PatternFill("solid", fgColor="CC0000")
        template["WK21"]["R14"] = "4 wekelijkse betaling"
        template["WK21"]["R14"].fill = PatternFill("solid", fgColor="CC0000")
        template["WK21"]["AA20"] = "=1+1"
        template["WK21"]["AA20"].fill = PatternFill("solid", fgColor="FFFF99")
        template["WK21"]["BB20"] = "oude losse template tekst"
        template["WK21"]["BB20"].fill = PatternFill("solid", fgColor="FFFF99")
        template["WK21"]["S130"] = "Loonbeslag Belastingdienst NL89"
        template["WK21"].sheet_view.topLeftCell = "R45"
        template["WK21"].sheet_view.selection[0].activeCell = "R45"
        template["WK21"].sheet_view.selection[0].sqref = "R45"
        template["Loonstrook"]["B7"] = "Werknemer"
        template["Loonstrook"]["B8"] = "=Periode!B8"
        template["Loonstrook"]["D8"] = "='WK21'!D8+'WK22'!D8+'WK23'!D8+'WK24'!D8"
        template["Loonstrook"]["P8"] = "oude loonstrooknotitie"
        template["Loonstrook"]["R30"] = "oude toelichting"

        period = {
            "workbook_tabs": [
                {
                    "label": "WK17",
                    "kind": "week",
                    "rows": [{
                        "employee_name": "Thomas",
                        "contract_hours": "40",
                        "worked_days": "5",
                        "worked_hours": "40",
                        "single_trip_km": "12",
                        "work_km": "4",
                        "total_km": "124",
                        "net_amount": "515,63",
                        "net_advance": "250",
                        "remarks": "dashboard opmerking",
                        "project_info": "Project A",
                    }],
                },
                {"label": "WK18", "kind": "week", "rows": []},
                {"label": "WK19", "kind": "week", "rows": []},
                {"label": "WK20", "kind": "week", "rows": []},
                {
                    "label": "Periode",
                    "kind": "period",
                    "rows": [{"employee_name": "Thomas", "contract_hours": "40", "gross_hourly_wage": "21,50", "net_period_basis": "650"}],
                },
                {"label": "Loonstrook", "kind": "payslip", "rows": [{"employee_name": "Thomas", "notes": "controle"}]},
            ]
        }
        with tempfile.TemporaryDirectory() as tmp_dir:
            template_path = Path(tmp_dir) / "template.xlsx"
            output_path = Path(tmp_dir) / "output.xlsx"
            template.save(template_path)
            build_tgn_template_output_workbook(output_path, period, template_path)
            workbook = load_workbook(output_path, data_only=False)

        self.assertIn("WK17", workbook.sheetnames)
        self.assertEqual(workbook["Periode"]["B8"].value, "Thomas")
        self.assertEqual(workbook["WK17"]["B8"].value, "Thomas")
        self.assertEqual(workbook["WK17"]["C8"].value, 40)
        self.assertEqual(workbook["WK17"]["E8"].value, 40)
        self.assertEqual(workbook["WK17"]["K8"].value, 515.63)
        self.assertEqual(workbook["WK17"]["L8"].value, 12)
        self.assertEqual(workbook["WK17"]["M8"].value, 4)
        self.assertEqual(workbook["WK17"]["N8"].value, 124)
        self.assertEqual(workbook["WK17"]["Q8"].value, 250)
        self.assertEqual(workbook["WK17"]["R8"].value, "dashboard opmerking")
        self.assertEqual(workbook["WK17"]["S8"].value, "Project A")
        self.assertIsNone(workbook["WK17"]["B2"].value)
        self.assertIsNone(workbook["WK17"]["B2"].fill.fill_type)
        self.assertIsNone(workbook["WK17"]["R14"].value)
        self.assertIsNone(workbook["WK17"]["R8"].fill.fill_type)
        self.assertIsNone(workbook["WK17"]["R14"].fill.fill_type)
        self.assertIsNone(workbook["WK17"]["AA20"].value)
        self.assertIsNone(workbook["WK17"]["AA20"].fill.fill_type)
        self.assertIsNone(workbook["WK17"]["BB20"].value)
        self.assertIsNone(workbook["WK17"]["BB20"].fill.fill_type)
        self.assertIsNone(workbook["WK17"]["S130"].value)
        self.assertEqual(workbook["Loonstrook"]["P8"].value, "controle")
        self.assertIsNone(workbook["Loonstrook"]["R30"].value)
        self.assertIn("'WK17'!D8", workbook["Loonstrook"]["D8"].value)
        self.assertEqual(workbook.active.title, "WK17")
        self.assertEqual(workbook["WK17"].sheet_view.topLeftCell, "A1")
        self.assertEqual(workbook["WK17"].sheet_view.selection[0].activeCell, "A1")

    @unittest.skipIf(Workbook is None, "openpyxl is niet beschikbaar")
    def test_tgn_template_export_leaves_empty_notes_empty(self):
        template = Workbook()
        template.active.title = "WK21"
        for sheet_name in ["WK22", "WK23", "WK24", "Periode", "Loonstrook", "Grondslag bouw & infra", "SAVG"]:
            template.create_sheet(sheet_name)
        template["Periode"]["B8"] = "Oud"
        template["WK21"]["B8"] = "=Periode!B8"
        template["WK21"]["R8"] = "oude opmerking"
        template["WK21"]["S8"] = "oude projectinfo"
        template["Loonstrook"]["P8"] = "oude loonstrooknotitie"
        period = {
            "workbook_tabs": [
                {"label": "WK17", "kind": "week", "rows": [{"employee_name": "Thomas", "worked_hours": "8"}]},
                {"label": "WK18", "kind": "week", "rows": []},
                {"label": "WK19", "kind": "week", "rows": []},
                {"label": "WK20", "kind": "week", "rows": []},
                {"label": "Periode", "kind": "period", "rows": [{"employee_name": "Thomas"}]},
                {"label": "Loonstrook", "kind": "payslip", "rows": [{"employee_name": "Thomas"}]},
            ]
        }
        with tempfile.TemporaryDirectory() as tmp_dir:
            template_path = Path(tmp_dir) / "template.xlsx"
            output_path = Path(tmp_dir) / "output.xlsx"
            template.save(template_path)
            build_tgn_template_output_workbook(output_path, period, template_path)
            workbook = load_workbook(output_path, data_only=False)

        self.assertIsNone(workbook["WK17"]["R8"].value)
        self.assertIsNone(workbook["WK17"]["S8"].value)
        self.assertIsNone(workbook["Loonstrook"]["P8"].value)

    def test_period_excel_endpoint_uses_tgn_template_export(self):
        source = Path("apps/dashboard/router.py").read_text(encoding="utf-8")
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8")
        requirements = Path("requirements.txt").read_text(encoding="utf-8-sig")

        self.assertIn("use_tgn_template=True", source)
        self.assertIn("TGN-templateopzet", source)
        self.assertIn("Excel TGN-opzet", template)
        self.assertIn("openpyxl==", requirements)

    def test_week_workbook_rows_link_to_timesheets(self):
        from apps.dashboard.payroll_calculations import WEEK_SHEET_COLUMNS, build_week_sheet_rows

        week = {"week_index": 1, "week_number": 18}
        rows = build_week_sheet_rows(
            "WK18",
            [{"id": 7, "name": "Thomas"}],
            [{"employee_name": "Thomas", "relation_id": 7, "week_hours": ["8", "0", "0", "0"], "week_timesheet_ids": [[44], [], [], []]}],
            week,
        )
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")

        self.assertIn({"label": "Urenbriefje", "key": "timesheet_link"}, WEEK_SHEET_COLUMNS)
        self.assertEqual(rows[0]["relation_id"], 7)
        self.assertEqual(rows[0]["timesheet_id"], 44)
        self.assertEqual(rows[0]["timesheet_link"], "Open")
        self.assertIn("payroll-workbook-link", template)
        self.assertIn("timesheet={{ row.timesheet_id }}", template)
        self.assertIn("/dashboard/relations?tab=candidates&edit={{ row.relation_id }}", template)

    def test_workbook_rows_do_not_invent_missing_payroll_values(self):
        from apps.dashboard.payroll_calculations import build_period_sheet_rows, build_week_sheet_rows

        period_rows = build_period_sheet_rows([], [{"employee_name": "Thomas"}])
        week_rows = build_week_sheet_rows("WK18", [], [{"employee_name": "Thomas", "week_hours": ["8"]}], {"week_index": 1, "week_number": 18})

        self.assertEqual(period_rows[0]["license_plate"], "")
        self.assertEqual(period_rows[0]["function_name"], "")
        self.assertEqual(period_rows[0]["gross_hourly_wage"], "")
        self.assertEqual(period_rows[0]["reserve_vacation_days"], "")
        self.assertEqual(week_rows[0]["single_trip_km"], "")
        self.assertEqual(week_rows[0]["project_info"], "")

    def test_workbook_editable_fields_recalculate_like_excel(self):
        week_row = {
            "worked_hours": "37.5",
            "worked_days": "5",
            "commute_km": "150",
            "work_km": "12,5",
            "single_trip_km": "15",
        }
        records._recalculate_payroll_derived_cells({"kind": "week"}, week_row)
        self.assertEqual(week_row["worked_hours"], "37.5")
        self.assertEqual(week_row["net_amount"], f"{chr(8364)} 515,63")
        self.assertEqual(week_row["total_km"], "162.5")

        period_row = {"contract_hours": "40", "gross_hourly_wage": f"{chr(8364)} 21,50"}
        records._recalculate_payroll_derived_cells({"kind": "period"}, period_row)
        self.assertEqual(period_row["gross_total"], f"{chr(8364)} 860,00")
        self.assertEqual(period_row["labor_cost_margin"], f"{chr(8364)} 154,80")

        script = Path("apps/dashboard/static/dashboard.js").read_text(encoding="utf-8-sig")
        self.assertIn("input.dataset.tabLabel.startsWith(\"WK\")", script)
        self.assertIn("gross-total", script)
        self.assertIn("normalized.includes(\",\") && normalized.includes(\".\")", script)



class PayrollRunningBalanceTests(unittest.TestCase):
    def test_running_balance_migration_tracks_required_balance_types(self):
        migration = Path("migrations/036_payroll_running_balances.sql").read_text(encoding="utf-8")

        self.assertIn("payroll_running_balance_accounts", migration)
        self.assertIn("payroll_running_balance_mutations", migration)
        self.assertIn("'wkr'", migration)
        self.assertIn("'loan_advance'", migration)
        self.assertIn("'choice_budget'", migration)
        self.assertIn("balance_year INTEGER NOT NULL DEFAULT 0", migration)

    def test_wkr_balance_status_warns_near_limit(self):
        self.assertEqual(records._running_balance_status("wkr", "2400", "2200"), "let op")
        self.assertEqual(records._running_balance_status("wkr", "2400", "2500"), "boven maximum")
        self.assertEqual(records._running_balance_status("loan_advance", None, "100"), "actief")



class PayrollPeriodSettlementTests(unittest.TestCase):
    def test_period_settlement_migration_creates_employee_period_totals(self):
        migration = Path("migrations/035_payroll_period_settlements.sql").read_text(encoding="utf-8")

        self.assertIn("payroll_period_settlements", migration)
        self.assertIn("advance_weeks_1_3", migration)
        self.assertIn("week_4_amount", migration)
        self.assertIn("total_period_amount", migration)

    def test_period_settlement_migration_supports_four_weekly_payment(self):
        migration = Path("migrations/035_payroll_period_settlements.sql").read_text(encoding="utf-8")

        self.assertIn("payment_schedule", migration)
        self.assertIn("four_weekly", migration)
        self.assertIn("total_period_amount", migration)



class PayrollWeekResultTests(unittest.TestCase):

    def test_employee_week_result_status_prioritizes_missing_data(self):
        self.assertEqual(records._employee_week_result_status(1, 1, 0), "mist inrichting")
        self.assertEqual(records._employee_week_result_status(1, 0, 1), "mist netto basisloon")
        self.assertEqual(records._employee_week_result_status(1, 0, 0), "concept")
        self.assertEqual(records._employee_week_result_status(0, 0, 0), "controle")
    def test_week_result_migration_creates_calculation_outputs(self):
        migration = Path("migrations/034_payroll_week_results.sql").read_text(encoding="utf-8")

        self.assertIn("payroll_week_results", migration)
        self.assertIn("net_wage_amount", migration)
        self.assertIn("travel_amount", migration)
        self.assertIn("day_allowance_amount", migration)
        self.assertIn("net_week_total", migration)

    def test_week_result_migration_tracks_missing_inputs(self):
        migration = Path("migrations/034_payroll_week_results.sql").read_text(encoding="utf-8")

        self.assertIn("mist_inrichting", migration)
        self.assertIn("mist_netto_basisloon", migration)
        self.assertIn("travel_km_net_uta", migration)
        self.assertIn("travel_km_net_build", migration)



class PayrollWeekInputTests(unittest.TestCase):
    def test_week_input_migration_creates_normalized_layers(self):
        migration = Path("migrations/033_payroll_week_inputs.sql").read_text(encoding="utf-8")

        self.assertIn("payroll_week_inputs", migration)
        self.assertIn("payroll_week_input_days", migration)
        self.assertIn("payroll_week_input_projects", migration)
        self.assertIn("idx_payroll_week_inputs_timesheet", migration)

    def test_week_input_days_cover_full_week(self):
        migration = Path("migrations/033_payroll_week_inputs.sql").read_text(encoding="utf-8")

        for day_name in ["maandag", "dinsdag", "woensdag", "donderdag", "vrijdag", "zaterdag", "zondag"]:
            self.assertIn(day_name, migration)

    def test_week_input_uses_calculated_km_when_total_is_missing(self):
        migration = Path("migrations/033_payroll_week_inputs.sql").read_text(encoding="utf-8")

        self.assertIn("calculated_total_km", migration)
        self.assertLess(
            migration.index("w.parsed_fields->'total_km'"),
            migration.index("w.parsed_fields->'calculated_total_km'"),
        )

    def test_parser_fills_total_km_from_day_km_when_total_missing(self):
        fields = {
            "monday_km": {"value": "34", "confidence": 98},
            "tuesday_km": {"value": "", "confidence": 0},
            "wednesday_km": {"value": "", "confidence": 0},
            "thursday_km": {"value": "", "confidence": 0},
            "friday_km": {"value": "", "confidence": 0},
            "saturday_km": {"value": "", "confidence": 0},
            "sunday_km": {"value": "", "confidence": 0},
            "total_km": {"value": "", "confidence": 0},
        }

        timesheet_parser._check_total_km(fields)

        self.assertEqual(fields["calculated_total_km"]["value"], "34")
        self.assertEqual(fields["total_km"]["value"], "34")
        self.assertEqual(fields["total_km_check"]["value"], "klopt")

    def test_corrections_materialize_missing_km_total_from_day_km(self):
        fields = {
            "monday_km": {"value": "34", "confidence": 98, "verified": True},
            "total_km": {"value": "", "confidence": 0},
        }

        timesheet_corrections._recalculate_total_checks(fields)

        self.assertEqual(fields["calculated_total_km"]["value"], "34")
        self.assertEqual(fields["total_km"]["value"], "34")
        self.assertEqual(fields["total_km_check"]["value"], "klopt")

    def test_payroll_rows_derive_days_and_km_from_parsed_day_fields(self):
        fields = {
            "monday_hours": {"value": "8"},
            "tuesday_hours": {"value": "8"},
            "wednesday_hours": {"value": "8"},
            "thursday_hours": {"value": "8"},
            "friday_hours": {"value": "8"},
            "saturday_hours": {"value": ""},
            "sunday_hours": {"value": ""},
            "monday_km": {"value": "21"},
            "tuesday_km": {"value": "21"},
            "wednesday_km": {"value": "21"},
            "thursday_km": {"value": "21"},
            "friday_km": {"value": "21"},
            "total_km": {"value": "105"},
            "total_hours": {"value": "40"},
        }

        self.assertEqual(records._parsed_worked_days(fields), records.Decimal("5"))
        self.assertEqual(records._parsed_total_hours(fields), records.Decimal("40"))
        self.assertEqual(records._parsed_total_km(fields), records.Decimal("105"))



class PayrollEmployeeArrangementTests(unittest.TestCase):
    def test_arrangement_migration_limits_period_numbers_to_thirteen(self):
        migration = Path("migrations/032_payroll_employee_arrangements.sql").read_text(encoding="utf-8")

        self.assertIn("valid_from_period_number BETWEEN 1 AND 13", migration)
        self.assertIn("payroll_employee_rights", migration)
        self.assertIn("payroll_employee_allowances", migration)

    def test_arrangement_payment_schedule_labels_are_documented(self):
        migration = Path("migrations/032_payroll_employee_arrangements.sql").read_text(encoding="utf-8")

        self.assertIn("'weekly'", migration)
        self.assertIn("'four_weekly'", migration)



class PayrollParameterTests(unittest.TestCase):
    def test_parameter_management_migration_is_applied(self):
        data_store = Path("apps/dashboard/data_store.py").read_text(encoding="utf-8")

        self.assertIn("048_payroll_parameter_management.sql", data_store)

    def test_parameter_management_seeds_blueprint_values(self):
        migration = Path("migrations/048_payroll_parameter_management.sql").read_text(encoding="utf-8")

        self.assertIn("vacation_days_18_plus", migration)
        self.assertIn("rv_days_build", migration)
        self.assertIn("training_reservation_percent", migration)
        self.assertIn("s3_shoes_annual", migration)
        self.assertIn("net_workable_days_per_year", migration)

    def test_settings_screen_can_save_parameter_versions(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8")

        self.assertIn('/api/settings/payroll-parameters', template)
        self.assertIn('name="parameter_version_id"', template)
        self.assertIn('name="effective_from"', template)
        self.assertIn('name="version_source_reference"', template)
        self.assertIn('name="notes"', template)
        self.assertIn('parameter_version={{ version.id }}', template)
        self.assertIn('@router.post("/api/settings/payroll-parameters")', router_source)
        self.assertIn('"payroll_parameter_version"', router_source)
        self.assertIn('"parameter_key"', router_source)

    def test_parameter_versions_keep_period_uniqueness(self):
        migration = Path("migrations/031_payroll_parameters.sql").read_text(encoding="utf-8")

        self.assertIn("UNIQUE (parameter_id, year, period_number)", migration)
        self.assertIn("period_number BETWEEN 1 AND 13", migration)

    def test_settings_cards_are_collapsible_by_default(self):
        script = Path("apps/dashboard/static/dashboard.js").read_text(encoding="utf-8")
        styles = Path("apps/dashboard/static/dashboard.css").read_text(encoding="utf-8")

        self.assertIn(".settings-only .settings-card", script)
        self.assertIn("settingsCollapsed", script)
        self.assertIn("activeSettingsTarget", script)
        self.assertIn('textContent = "Openen"', script)
        self.assertIn('[data-settings-collapsed="true"]', styles)

    def test_formats_parameter_percentages_for_display(self):
        self.assertEqual(records._format_parameter_value("0.0833", "percentage"), "8.33%")
        self.assertEqual(records._format_parameter_value("0.0800", "percentage"), "8%")

    def test_formats_parameter_money_for_display(self):
        self.assertIn("0,28", records._format_parameter_value("0.28", "euro_per_km"))



class PayrollPhaseStatusTests(unittest.TestCase):
    def test_phase_status_blocks_without_week_results(self):
        status = records.payroll_phase_status({"result_count": 0}, {"blocking": 0, "warning": 0, "total": 0})

        self.assertFalse(status["can_approve"])
        self.assertEqual(status["label"], "Nog niet berekend")
        self.assertIn("geen gevalideerde uren", status["detail"])

    def test_phase_status_blocks_known_exceptions(self):
        status = records.payroll_phase_status({"result_count": 4}, {"blocking": 1, "warning": 2, "total": 3})

        self.assertFalse(status["can_approve"])
        self.assertEqual(status["tone"], "danger")

    def test_phase_status_allows_approval_with_warnings_only(self):
        status = records.payroll_phase_status({"result_count": 4}, {"blocking": 0, "warning": 1, "total": 1})

        self.assertTrue(status["can_approve"])
        self.assertEqual(status["label"], "Nalopen voor akkoord")


class PayrollExceptionTests(unittest.TestCase):
    def test_exception_summary_counts_severities(self):
        summary = records.summarize_payroll_exceptions([
            {"severity": "blokkerend"},
            {"severity": "waarschuwing"},
            {"severity": "waarschuwing"},
            {"severity": "info"},
        ])

        self.assertEqual(summary, {"total": 4, "blocking": 1, "warning": 2, "info": 1})

    def test_exception_severity_labels_match_period_ui(self):
        self.assertEqual(records._payroll_exception_severity_label("blokkerend"), "Blokkeert")
        self.assertEqual(records._payroll_exception_severity_label("waarschuwing"), "Nalopen")
        self.assertEqual(records._payroll_exception_severity_label("anders"), "Info")

    def test_partial_period_is_not_a_payroll_exception(self):
        source = inspect.getsource(records.list_payroll_period_exceptions)

        self.assertNotIn("incomplete_period", source)
        self.assertNotIn("Periode niet compleet", source)


class RelationPayrollContextTests(unittest.TestCase):
    def test_relation_payroll_context_combines_employee_layers(self):
        arrangement = {"id": 11, "relation_id": 7}
        balance = {"id": 21, "relation_id": 7}
        settlement = {"period_label": "2026 P5", "relation_id": 7}

        with patch.object(records, "list_relation_payroll_employee_arrangements", return_value=[arrangement]), \
             patch.object(records, "list_relation_payroll_running_balances", return_value=[balance]), \
             patch.object(records, "list_relation_payroll_period_settlements", return_value=[settlement]):
            context = records.get_relation_payroll_context(7)

        self.assertEqual(context["current_arrangement"], arrangement)
        self.assertEqual(context["balances"], [balance])
        self.assertEqual(context["settlements"], [settlement])

    def test_relation_payroll_context_handles_empty_relation(self):
        self.assertEqual(
            records.get_relation_payroll_context(None),
            {"arrangements": [], "current_arrangement": None, "balances": [], "settlements": []},
        )


class PayrollPeriodStructureTests(unittest.TestCase):
    def test_available_payroll_period_numbers_stop_at_thirteen(self):
        with patch.object(records, "ensure_dashboard_tables", side_effect=RuntimeError("geen database")):
            numbers = records._available_payroll_period_numbers(2026, 20)

        self.assertEqual(numbers, list(range(1, records.PAYROLL_PERIODS_PER_YEAR + 1)))

    def test_create_payroll_period_rejects_period_fourteen(self):
        with patch.object(records, "ensure_dashboard_tables"):
            with self.assertRaises(ValueError):
                records.create_payroll_period({"year": "2026", "period_number": "14"})

    def test_payroll_period_display_numbers_remain_real_period_numbers(self):
        periods = [
            {
                "id": 5,
                "year": 2026,
                "period_number": 5,
                "start_date": "22-06-2026",
                "end_date": "19-07-2026",
                "name": "Periode 05 22/06 - 19/07",
            }
        ]

        records._apply_period_display_numbers(periods)

        self.assertEqual(periods[0]["display_period_number"], 5)
        self.assertEqual(periods[0]["name"], "Periode 05 22/06 - 19/07")

    def test_payroll_periods_can_only_be_archived_from_dashboard(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn("Archiveer", template)
        self.assertIn("Terugzetten", template)
        self.assertNotIn("Verwijder definitief", template)
        self.assertNotIn("delete_payroll_period,", router_source)
        self.assertIn("Verwijderen is uitgeschakeld", router_source)
        delete_block = records_source[records_source.index("def delete_payroll_period"):records_source.index("def update_payroll_period_status")]
        self.assertIn("archive_payroll_period(period_id, archived=True)", delete_block)
        self.assertNotIn("DELETE FROM payroll_periods", delete_block)


class PayrollDatamodelFoundationTests(unittest.TestCase):
    def test_foundation_migration_creates_year_and_week_line_tables(self):
        migration = Path("migrations/037_payroll_datamodel_foundation.sql").read_text(encoding="utf-8-sig")

        self.assertIn("payroll_years", migration)
        self.assertIn("period_count INTEGER NOT NULL DEFAULT 13", migration)
        self.assertIn("weeks_per_period INTEGER NOT NULL DEFAULT 4", migration)
        self.assertIn("payroll_week_lines", migration)
        self.assertIn("cost_center", migration)
        self.assertIn("UNIQUE (payroll_week_input_id, line_index)", migration)

    def test_foundation_migration_extends_openai_audit_for_ocr_context(self):
        migration = Path("migrations/037_payroll_datamodel_foundation.sql").read_text(encoding="utf-8-sig")

        for field in [
            "purpose",
            "relation_id",
            "timesheet_inbox_id",
            "payroll_week_input_id",
            "request_hash",
            "response_hash",
            "prompt_tokens",
            "completion_tokens",
            "total_tokens",
            "latency_ms",
        ]:
            self.assertIn(field, migration)

    def test_foundation_migration_is_part_of_dashboard_startup(self):
        data_store = Path("apps/dashboard/data_store.py").read_text(encoding="utf-8-sig")

        self.assertIn("migrations/037_payroll_datamodel_foundation.sql", data_store)



class PayrollAuditContextTests(unittest.TestCase):
    def test_audit_context_migration_links_payroll_and_ai_ocr_records(self):
        migration = Path("migrations/042_payroll_audit_context.sql").read_text(encoding="utf-8-sig")

        for field in [
            "payroll_year_id",
            "payroll_period_id",
            "payroll_period_week_id",
            "payroll_week_input_id",
            "timesheet_inbox_id",
            "relation_id",
            "correlation_id",
            "source_channel",
        ]:
            self.assertIn(field, migration)
        self.assertIn("payroll_audit_context", migration)
        self.assertIn("payroll_ai_ocr_audit_context", migration)
        self.assertIn("payroll_period_audit_summary", migration)

    def test_audit_context_migration_is_part_of_dashboard_startup(self):
        data_store = Path("apps/dashboard/data_store.py").read_text(encoding="utf-8-sig")

        self.assertIn("migrations/042_payroll_audit_context.sql", data_store)
        self.assertLess(
            data_store.index("migrations/038_payroll_datamodel_views.sql"),
            data_store.index("migrations/042_payroll_audit_context.sql"),
        )

    def test_audit_context_fields_are_inferred_from_entity_and_metadata(self):
        context = records._audit_context_fields(
            "payroll_period",
            12,
            {"relation_id": "7", "timesheet_inbox_id": "44", "source_channel": "test"},
        )

        self.assertEqual(context["payroll_period_id"], 12)
        self.assertEqual(context["relation_id"], 7)
        self.assertEqual(context["timesheet_inbox_id"], 44)
        self.assertEqual(context["source_channel"], "test")

    def test_openai_audit_context_defaults_to_timesheet_ocr(self):
        context = openai_usage._openai_api_audit_context(
            "whatsapp_timesheet",
            55,
            {"relation_id": "8", "total_tokens": "120"},
        )

        self.assertEqual(context["provider"], "openai")
        self.assertEqual(context["purpose"], "timesheet_ocr")
        self.assertEqual(context["timesheet_inbox_id"], 55)
        self.assertEqual(context["relation_id"], 8)
        self.assertEqual(context["total_tokens"], 120)


class PayrollDatamodelViewTests(unittest.TestCase):
    def test_datamodel_views_migration_creates_period_and_year_views(self):
        migration = Path("migrations/038_payroll_datamodel_views.sql").read_text(encoding="utf-8-sig")

        self.assertIn("payroll_year_overview", migration)
        self.assertIn("payroll_period_datamodel_status", migration)
        self.assertIn("expected_period_count", migration)
        self.assertIn("actual_week_count", migration)
        self.assertIn("week_structure_status", migration)

    def test_period_datamodel_status_tracks_foundation_layers(self):
        migration = Path("migrations/038_payroll_datamodel_views.sql").read_text(encoding="utf-8-sig")

        for field in [
            "week_input_count",
            "week_line_count",
            "week_result_count",
            "period_settlement_count",
            "employee_arrangement_count",
            "parameter_version_count",
            "running_balance_account_count",
            "audit_event_count",
            "openai_api_audit_event_count",
        ]:
            self.assertIn(field, migration)

    def test_datamodel_views_migration_is_part_of_dashboard_startup(self):
        data_store = Path("apps/dashboard/data_store.py").read_text(encoding="utf-8-sig")

        self.assertIn("migrations/038_payroll_datamodel_views.sql", data_store)


class PayrollDatamodelRecordTests(unittest.TestCase):
    def test_datamodel_status_row_formats_counts_and_dates(self):
        row = (
            12,
            2026,
            5,
            "Periode 5 - 2026",
            date(2026, 5, 4),
            date(2026, 5, 31),
            "open",
            4,
            10,
            14,
            8,
            3,
            7,
            16,
            21,
            2,
            5,
            6,
            "ok",
            datetime(2026, 6, 17, 9, 30),
        )

        item = records._payroll_datamodel_status_row(row)

        self.assertEqual(item["period_number"], 5)
        self.assertEqual(item["week_count"], 4)
        self.assertEqual(item["week_line_count"], 14)
        self.assertEqual(item["parameter_version_count"], 16)
        self.assertEqual(item["week_structure_status"], "ok")
        self.assertEqual(item["start_date"], "04-05-2026")

    def test_year_overview_row_formats_expected_structure(self):
        row = (1, 2026, 13, 4, 2, 8, date(2026, 1, 5), date(2026, 2, 28), "active", None)

        item = records._payroll_year_overview_row(row)

        self.assertEqual(item["expected_period_count"], 13)
        self.assertEqual(item["expected_weeks_per_period"], 4)
        self.assertEqual(item["actual_week_count"], 8)
        self.assertEqual(item["status"], "active")


class PayrollDatamodelDashboardTests(unittest.TestCase):
    def test_periods_page_loads_datamodel_context(self):
        router = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")

        self.assertIn("list_payroll_year_overview", router)
        self.assertIn("list_payroll_datamodel_status", router)
        self.assertIn("payroll_year_overview", router)
        self.assertIn("payroll_datamodel_status", router)

    def test_periods_template_shows_datamodel_control_section(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        stylesheet = Path("apps/dashboard/static/dashboard.css").read_text(encoding="utf-8-sig")

        self.assertIn("datamodel-controle", template)
        self.assertIn("Fundament jaar, periodes en payroll-lagen", template)
        self.assertIn("week_line_count", template)
        self.assertIn("openai_api_audit_event_count", template)
        self.assertIn(".datamodel-check", stylesheet)

class PayrollPeriodDatamodelDetailTests(unittest.TestCase):
    def test_period_detail_attaches_datamodel_status(self):
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn("def get_payroll_period_datamodel_status", records_source)
        self.assertIn('period["datamodel_status"] = get_payroll_period_datamodel_status(period_id)', records_source)
        self.assertIn("FROM payroll_period_datamodel_status", records_source)

    def test_period_detail_template_shows_datamodel_strip(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        stylesheet = Path("apps/dashboard/static/dashboard.css").read_text(encoding="utf-8-sig")

        self.assertIn("period-datamodel-strip", template)
        self.assertIn("selected_payroll_period.datamodel_status", template)
        self.assertIn("week_line_count", template)
        self.assertIn(".period-datamodel-strip", stylesheet)

class PayrollEmptyPeriodCopyTests(unittest.TestCase):
    def test_empty_period_copy_explains_missing_validated_hours(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")

        self.assertIn("Nog geen gevalideerde uren", template)
        self.assertIn("Valideer eerst urenbriefjes", template)
        self.assertIn("gevalideerde uren in deze loonperiode", template)


class PayrollTestDataMigrationSafetyTests(unittest.TestCase):
    def test_period_two_testdata_migration_preserves_validated_statuses(self):
        migration = Path("migrations/028_payroll_period_02_test_timesheets.sql").read_text(encoding="utf-8-sig")

        self.assertIn("LOWER(REPLACE(COALESCE(w.status, ''), ' ', '_')) IN ('', 'controle', 'te_controleren', 'gematcht', 'matched')", migration)
        self.assertGreaterEqual(
            migration.count("LOWER(REPLACE(COALESCE(w.status, ''), ' ', '_')) IN ('', 'controle', 'te_controleren', 'gematcht', 'matched')"),
            4,
        )
        self.assertIn("ELSE w.status", migration)
        self.assertIn("ELSE w.validated_at", migration)
        self.assertIn("ELSE w.payroll_sent_at", migration)
        self.assertNotIn("SET status = 'controle',", migration)


class DashboardOverviewWeeklyHoursTests(unittest.TestCase):
    def test_demo_weekly_hours_are_marked_and_link_to_periods(self):
        rows = records._demo_weekly_hours_yoy()

        self.assertTrue(rows)
        self.assertTrue(all(row["is_demo"] for row in rows))
        self.assertTrue(all(row["source_label"] == "Demo" for row in rows))
        self.assertTrue(all(row["href"] == "/dashboard/periods#periodes" for row in rows))

    def test_weekly_hours_cards_are_clickable(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")

        self.assertIn("href=\"{{ week.href|default('/dashboard/periods#periodes') }}\"", template)
        self.assertIn("hours-yoy-card--demo", template)
        self.assertIn("voorbeelddata", template)


class DashboardVisibleDemoPayrollSeedTests(unittest.TestCase):
    def test_empty_dashboard_views_trigger_visible_demo_seed(self):
        source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn("def _ensure_dashboard_tables_for_read", source)
        self.assertIn("DASHBOARD_SCHEMA_ENSURE_READ_WARNING", source)
        self.assertIn("def ensure_visible_demo_payroll_data", source)
        self.assertIn("migrations/039_full_year_test_payroll.sql", source)
        self.assertIn("migrations/041_dashboard_demo_payroll.sql", source)
        self.assertIn("migrations/033_payroll_week_inputs.sql", source)
        self.assertIn("DASHBOARD_DEMO_PAYROLL_SEED_STEP_ERROR", source)
        runtime_source = source[source.index("def get_overview_data"):]
        self.assertNotIn("ensure_visible_demo_payroll_data()", runtime_source)
        self.assertGreaterEqual(source.count("_ensure_dashboard_tables_for_read()"), 10)

    def test_payroll_archive_is_a_real_tab_panel(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        script = Path("apps/dashboard/static/dashboard.js").read_text(encoding="utf-8-sig")

        self.assertIn("data-period-tabs", template)
        self.assertIn("data-period-tab=\"archive\"", template)
        self.assertIn("data-period-panel=\"archive\"", template)
        self.assertIn("hidden", template)
        self.assertIn("data-period-panel", script)
        self.assertIn("#periode-archief", script)

    def test_period_page_ensures_full_2026_calendar_without_timesheets(self):
        source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn("def ensure_payroll_period_calendar", source)
        self.assertIn("PAYROLL_CALENDAR_START_2026", source)
        self.assertIn("range(1, PAYROLL_PERIODS_PER_YEAR + 1)", source)
        self.assertIn("INSERT INTO payroll_periods", source)
        self.assertIn("INSERT INTO payroll_period_weeks", source)
        self.assertIn("ensure_payroll_period_calendar(2026)", source)
        list_periods_block = source[source.index("def list_payroll_periods"):source.index("def get_payroll_data_diagnostics")]
        self.assertNotIn("ensure_visible_demo_payroll_data()", list_periods_block)

class PayrollDataDiagnosticsTests(unittest.TestCase):
    def test_periods_page_loads_payroll_data_diagnostics(self):
        router = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn("def get_payroll_data_diagnostics", records_source)
        self.assertIn("get_payroll_data_diagnostics", router)
        self.assertIn("payroll_data_diagnostics", router)
        self.assertIn("payroll-data-controle", template)
        self.assertIn("Echte tabeldata", template)
        self.assertIn("Projectboekingen", records_source)
        self.assertIn("AI/OCR audit", records_source)


class CompletePeriodTimesheetImportTests(unittest.TestCase):
    def test_zip_import_iterates_supported_documents_only(self):
        buffer = BytesIO()
        with zipfile.ZipFile(buffer, "w") as archive:
            archive.writestr("deel/A weekstaat 21 2026.jpeg", b"image")
            archive.writestr("deel/B weekstaat 21 2026.pdf", b"pdf")
            archive.writestr("deel/notitie.docx", b"docx")

        docs = list(timesheet_uploads._iter_import_documents("periode.zip", buffer.getvalue()))

        self.assertEqual(len(docs), 2)
        self.assertTrue(docs[0][0].startswith("periode/"))
        self.assertTrue(any(name.endswith(".pdf") for name, _ in docs))

    def test_complete_period_import_ui_and_route_are_present(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        upload_source = Path("apps/dashboard/timesheet_uploads.py").read_text(encoding="utf-8-sig")

        self.assertIn("/api/whatsapp/complete-period-import", template)
        self.assertIn("Upload testset", template)
        self.assertIn("complete-period-import", router_source)
        self.assertIn("replace_complete_period_import", upload_source)
        self.assertIn("project_time_bookings", upload_source)

    def test_complete_period_replace_clears_old_payroll_processing(self):
        upload_source = Path("apps/dashboard/timesheet_uploads.py").read_text(encoding="utf-8-sig")
        replace_block = upload_source[upload_source.index("def replace_complete_period_import"):upload_source.index("def import_complete_period_timesheets")]

        self.assertIn("SELECT id", replace_block)
        self.assertIn("payroll_week_inputs", replace_block)
        self.assertIn("payroll_week_results", replace_block)
        self.assertIn("payroll_week_lines", replace_block)
        self.assertIn("payroll_period_settlements", replace_block)
        self.assertIn("payroll_period_totals", replace_block)
        self.assertIn("DELETE FROM whatsapp_timesheet_inbox", replace_block)
        self.assertIn("ensure_payroll_period_calendar(2026)", replace_block)

    def test_timesheet_upload_falls_back_when_candidate_fk_is_stale(self):
        upload_source = Path("apps/dashboard/timesheet_uploads.py").read_text(encoding="utf-8-sig")

        self.assertIn("from psycopg2.errors import ForeignKeyViolation", upload_source)
        insert_block = upload_source.split("INSERT INTO whatsapp_timesheet_inbox", 1)[1].split("RETURNING id", 1)[0]
        self.assertNotIn("matched_candidate_id", insert_block)
        self.assertIn("except ForeignKeyViolation", upload_source)
        self.assertIn('status = "te_controleren"', upload_source)

    def test_legacy_timesheet_candidate_fk_is_removed_by_migration(self):
        data_store = Path("apps/dashboard/data_store.py").read_text(encoding="utf-8-sig")
        migration = Path("migrations/046_clear_legacy_timesheet_candidate_fk.sql").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")

        self.assertIn("046_clear_legacy_timesheet_candidate_fk.sql", data_store)
        self.assertIn("whatsapp_timesheet_inbox", migration)
        self.assertIn("matched_candidate_id", migration)
        self.assertIn("DROP CONSTRAINT", migration)
        self.assertIn("_db_error_detail", router_source)
        self.assertIn("constraint_name", router_source)

    def test_payroll_test_workspace_can_be_cleared_from_ui(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn("/api/test/payroll-workspace/clear", template)
        self.assertIn("Reset testdata", template)
        self.assertIn("def clear_payroll_workspace_for_testing", router_source)
        self.assertIn("clear_payroll_test_workspace", router_source)
        self.assertNotIn("cleared_periods", router_source)
        self.assertIn("def _delete_all_existing_tables", records_source)
        clear_block = records_source[records_source.index("def clear_payroll_test_workspace"):records_source.index("def archive_payroll_period")]
        self.assertNotIn("TRUNCATE TABLE", clear_block)
        self.assertIn('DELETE FROM "{table_name}"', records_source)
        self.assertIn("project_time_bookings", records_source)
        self.assertIn("whatsapp_timesheet_inbox", records_source)
        self.assertIn("Loonperiodes zijn behouden", records_source)
        self.assertIn("ensure_payroll_period_calendar(2026)", clear_block)
        self.assertIn("payroll_week_inputs", records_source)
        self.assertIn("payroll_week_results", records_source)
        self.assertIn("payroll_period_settlements", records_source)
        self.assertIn("openai_api_audit_events", records_source)
        self.assertIn("audit_events", records_source)
        self.assertIn("_payroll_demo_seed_is_suppressed", records_source)
        self.assertIn("cleared_payroll_rows", router_source)

    def test_test_data_controls_are_visually_grouped(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        stylesheet = Path("apps/dashboard/static/dashboard.css").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")

        self.assertIn("test-data-panel", template)
        self.assertIn("Upload testset", template)
        self.assertIn("test-file-input", template)
        self.assertIn("test-action-button--primary", template)
        self.assertIn("test-action-button--danger", template)
        self.assertIn("Reset testdata", template)
        self.assertIn("De loonperiodes zelf blijven staan", template)
        self.assertNotIn("loonperiodes voor deze testomgeving legen", template)
        self.assertIn(".test-data-panel", stylesheet)
        self.assertIn(".test-file-input", stylesheet)
        self.assertIn("@router.get(\"/api/whatsapp/complete-period-import\")", router_source)
        self.assertIn("upload_error", router_source)
        self.assertIn("import_error", template)

    def test_test_seeds_are_suppressed_after_reset_and_deploy_clears_once(self):
        data_store = Path("apps/dashboard/data_store.py").read_text(encoding="utf-8-sig")
        migration = Path("migrations/043_reset_payroll_test_dataset_once.sql").read_text(encoding="utf-8-sig")
        fast_migration = Path("migrations/044_fast_reset_payroll_test_dataset_once.sql").read_text(encoding="utf-8-sig")
        restore_migration = Path("migrations/045_restore_payroll_period_calendar.sql").read_text(encoding="utf-8-sig")

        self.assertIn("_payroll_test_seed_is_suppressed", data_store)
        self.assertLess(
            data_store.index("migrations/021_audit_events.sql"),
            data_store.index("migrations/022_otys_staging_tables.sql"),
        )
        self.assertNotIn("043_reset_payroll_test_dataset_once.sql", data_store)
        self.assertNotIn("044_fast_reset_payroll_test_dataset_once.sql", data_store)
        self.assertIn("045_restore_payroll_period_calendar.sql", data_store)
        self.assertNotIn("migrations/019_demo_seed_data.sql", data_store)
        self.assertNotIn("migrations/040_one_period_test_hours.sql", data_store)
        self.assertIn("deploy_reset_043", migration)
        self.assertIn("deploy_reset_044", fast_migration)
        self.assertIn("TRUNCATE TABLE", migration)
        self.assertIn("TRUNCATE TABLE", fast_migration)
        self.assertIn("whatsapp_timesheet_inbox", migration)
        self.assertIn("payroll_periods", migration)
        self.assertIn("openai_api_audit_events", migration)
        self.assertIn("Testfase uren en loonperiodes geleegd", migration)
        self.assertIn("Herstelde loonperiodekalender", restore_migration)
        self.assertIn("generate_series(1, 13)", restore_migration)
        self.assertIn("payroll_period_weeks", restore_migration)

    def test_audit_context_migration_ignores_deleted_test_records(self):
        migration = Path("migrations/042_payroll_audit_context.sql").read_text(encoding="utf-8-sig")

        self.assertIn("FROM whatsapp_timesheet_inbox w", migration)
        self.assertIn("WHERE w.id = e.entity_id", migration)
        self.assertIn("WHERE w.id = a.source_id", migration)
        self.assertIn("FROM relations r", migration)
        self.assertIn("FROM payroll_periods p", migration)
        self.assertIn("FROM payroll_week_inputs i", migration)

    def test_payslip_manual_net_received_recalculates_remaining_net(self):
        row = {
            "period_total": f"{chr(8364)} 1707,06",
            "already_received_net": "500",
            "payslip_advance": f"{chr(8364)} 0,00",
            "net_to_receive": f"{chr(8364)} 1707,06",
            "net_total": f"{chr(8364)} 1707,06",
        }

        records._recalculate_payroll_derived_cells({"kind": "payslip"}, row)

        self.assertEqual(row["net_to_receive"], f"{chr(8364)} 1.207,06")
        self.assertEqual(row["net_total"], f"{chr(8364)} 1.207,06")

    def test_period_settlement_table_does_not_show_implicit_advance_split(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")

        self.assertNotIn("Voorschot wk 1-3", template)
        self.assertNotIn("Week 4 / uitbetaling", template)
        self.assertNotIn("result.advance_weeks_1_3", template)
        self.assertNotIn("result.week_4_amount", template)

    def test_period_payroll_rows_are_sourced_from_week_inputs(self):
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")
        block = records_source[
            records_source.index("def list_payroll_period_payroll"):
            records_source.index("_PAYROLL_DAY_KEYS")
        ]

        self.assertIn("FROM payroll_week_inputs i", block)
        self.assertIn("i.payroll_period_id = %s", block)
        self.assertIn("i.payroll_period_week_id", block)
        self.assertIn("_active_period_payroll_status_condition(\"i\", \"pp\")", block)
        self.assertNotIn("FROM whatsapp_timesheet_inbox w", block)

    def test_payroll_payment_flow_has_workbook_tabs_and_actions(self):
        calculations_source = Path("apps/dashboard/payroll_calculations.py").read_text(encoding="utf-8-sig")
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn('"label": "Uit te betalen"', calculations_source)
        self.assertIn('"label": "Uitbetaald"', calculations_source)
        self.assertIn('"key": "payment_action"', calculations_source)
        payment_columns = calculations_source[
            calculations_source.index("PAYMENT_SHEET_COLUMNS = ["):
            calculations_source.index("DEFAULT_RULES = [")
        ]
        self.assertNotIn('"key": "payroll_status_label"', payment_columns)
        self.assertIn("/api/periods/{{ selected_payroll_period.id }}/payment-status", template)
        self.assertIn("Uitbetalen", template)
        self.assertIn("Uitbetaald", template)
        self.assertIn("def save_payroll_period_payment_status", router_source)
        self.assertIn("def update_payroll_payment_status", records_source)
        self.assertIn('"uit_te_betalen"', records_source)
        self.assertIn('"uitbetaald"', records_source)
        self.assertIn("COALESCE(i.status, pc.booking_status, 'concept') AS status", records_source)
        self.assertNotIn("COALESCE(pc.booking_status, i.status) AS status", records_source)
        self.assertIn("payment_source_tabs", calculations_source)
        self.assertIn('row.get("payroll_status") == "loon_berekenen"', calculations_source)
        self.assertIn('target_tab = "Uitbetaald" if result.get("status") == "uitbetaald" else "Uit te betalen"', router_source)
        self.assertIn("PAYROLL_PREPAYMENT_STATUSES", records_source)
        self.assertIn("def _active_timesheet_condition", records_source)
        self.assertIn("payroll-week-tab-summary", template)
        self.assertIn("uur - {{ tab.summary.booking_count }} regels", template)
        self.assertIn("def summarize_payroll_payment_flow", records_source)
        self.assertIn('period["payroll_payment_summary"] = summarize_payroll_payment_flow(period["workbook_tabs"])', records_source)
        self.assertIn("Uit te betalen loon", template)
        self.assertIn("Uitbetaald loon", template)
        self.assertIn("Open loonregels", template)
        self.assertIn("Blokkades", template)

    def test_deleted_timesheets_are_removed_from_payroll_periods(self):
        actions_source = Path("apps/dashboard/whatsapp_actions.py").read_text(encoding="utf-8-sig")
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn("DELETE FROM payroll_week_inputs", actions_source)
        self.assertIn("WHERE timesheet_inbox_id = %s", actions_source)
        self.assertIn("LEFT JOIN whatsapp_timesheet_inbox wi", records_source)
        self.assertIn("_active_timesheet_condition(\"i\", \"wi\")", records_source)


class DashboardDatabaseStatusTests(unittest.TestCase):
    def test_non_overview_pages_use_real_database_status(self):
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")

        self.assertIn("get_database_status", router_source)
        self.assertIn('"database": database_status', router_source)
        self.assertIn('"database_status"', router_source)

    def test_relation_detail_read_continues_after_schema_warning(self):
        relation_source = Path("apps/dashboard/relations.py").read_text(encoding="utf-8-sig")

        self.assertIn("RELATION_READ_SCHEMA_WARNING", relation_source)
        self.assertIn("SELECT * FROM relations WHERE id = %s", relation_source)


class TimesheetCandidateValidationTests(unittest.TestCase):
    def test_timesheet_validation_requires_candidate_and_inline_creation(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        corrections_source = Path("apps/dashboard/timesheet_corrections.py").read_text(encoding="utf-8-sig")
        stylesheet = Path("apps/dashboard/static/dashboard.css").read_text(encoding="utf-8-sig")

        self.assertIn("Koppel eerst een kandidaat", template)
        self.assertIn("Kandidaat aanmaken", template)
        self.assertIn("data-workflow-candidate-id", template)
        self.assertIn("data-workflow-validate-button", template)
        self.assertIn("/api/whatsapp/timesheet/{{ selected_message.id }}/candidate", template)
        self.assertIn("timesheet-create-candidate-form", template)
        self.assertIn("@router.post(\"/api/whatsapp/timesheet/{timesheet_id}/candidate\")", router_source)
        self.assertIn("create_candidate(", router_source)
        self.assertIn("save_field_corrections(", router_source)
        self.assertIn("TimesheetValidationError", corrections_source)
        self.assertIn("Koppel eerst een kandidaat", corrections_source)
        self.assertIn("relation_type = 'candidate'", corrections_source)
        self.assertIn("validate_error", router_source)
        self.assertIn("matched_relation_id: str = Form(\"\")", router_source)
        self.assertIn(".candidate-create-panel", stylesheet)
        self.assertIn(".workflow-action-note", stylesheet)

    def test_candidate_selection_enables_validation_button(self):
        script = Path("apps/dashboard/static/dashboard.js").read_text(encoding="utf-8-sig")

        self.assertIn("data-workflow-candidate-id", script)
        self.assertIn("data-workflow-validate-button", script)
        self.assertIn('canApproveControl = validateButton?.dataset.canApproveControl === "1"', script)
        self.assertIn('canSendToPayroll = validateButton?.dataset.canSendToPayroll === "1"', script)
        self.assertIn("canUseWorkflowButton = canApproveControl || canSendToPayroll", script)
        self.assertIn("validateButton.disabled = !canUseWorkflowButton", script)
        self.assertIn("workflowCandidateTarget.value = option.value", script)

    def test_control_stage_can_advance_to_validation_without_payroll(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        correction_source = Path("apps/dashboard/timesheet_corrections.py").read_text(encoding="utf-8-sig")

        self.assertIn("approve-control", template)
        self.assertIn("Naar valideren", template)
        self.assertIn("data-can-approve-control", template)
        self.assertIn('@router.post("/api/whatsapp/timesheet/{timesheet_id}/approve-control")', router_source)
        self.assertIn("approve_timesheet_control", router_source)
        self.assertIn("Controle urenbriefje akkoord", router_source)
        self.assertIn("def approve_timesheet_control", correction_source)
        self.assertIn("status = 'goed_te_keuren'", correction_source)

    def test_failed_parse_can_be_completed_manually_with_audit_marker(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        script = Path("apps/dashboard/static/dashboard.js").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        stylesheet = Path("apps/dashboard/static/dashboard.css").read_text(encoding="utf-8-sig")

        self.assertIn("OCR + OpenAI parsen", template)
        self.assertIn("form=\"timesheet-force-parse-form\"", template)
        self.assertIn("/api/whatsapp/timesheet/{{ selected_message.id }}/reparse", template)
        self.assertIn("data-force-parse-button", template)
        self.assertIn("data-manual-fields-open", template)
        self.assertIn('name="manual_parse"', template)
        self.assertIn("Handmatige parse opslaan", template)
        self.assertIn("form.querySelectorAll(\".timesheet-accordion\")", script)
        self.assertIn("OCR + OpenAI bezig", script)
        self.assertIn("formData.set(submitter.name", script)
        self.assertIn("manual_parse = str(form.get(\"manual_parse\")", router_source)
        self.assertIn("Urenbriefje handmatig geparsed", router_source)
        self.assertIn("Urenbriefje met OCR en OpenAI geparsed", router_source)
        self.assertIn("tab=task&stage=controle", router_source)
        self.assertIn(".manual-parse-actions", stylesheet)

    def test_manual_timesheet_fields_recalculate_totals_and_badges(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        script = Path("apps/dashboard/static/dashboard.js").read_text(encoding="utf-8-sig")
        parser_source = Path("apps/dashboard/timesheet_parser.py").read_text(encoding="utf-8-sig")
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn('data-manual-parse-save="true"', template)
        self.assertIn("'week_number')", template)
        self.assertIn('type="date" name="field_date"', template)
        self.assertIn("selected_message.date_input_value", template)
        self.assertIn("_timesheet_date_input_value", records_source)
        self.assertIn('"%d-%m-%y"', parser_source)
        self.assertIn("'principal_name')", template)
        self.assertIn("forceTotals || source === \"hours\"", script)
        self.assertIn("forceTotals || source === \"km\"", script)
        self.assertIn("totalField.value = calculatedField.value", script)
        self.assertIn("syncEditableSummary", script)

    def test_payroll_employee_settings_and_balances_are_editable_with_audit(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")
        script = Path("apps/dashboard/static/dashboard.js").read_text(encoding="utf-8-sig")
        stylesheet = Path("apps/dashboard/static/dashboard.css").read_text(encoding="utf-8-sig")

        self.assertIn("/api/payroll/employee-arrangements/{{ arrangement.id }}", template)
        self.assertIn("/api/payroll/running-balances/{{ balance.id }}", template)
        self.assertIn("/api/payroll/running-balances/{{ balance.id }}/mutations", template)
        self.assertIn("data-settings-edit-toggle", template)
        self.assertIn("settings-inline-form", template)
        self.assertIn("update_payroll_employee_arrangement", router_source)
        self.assertIn("update_payroll_running_balance_account", router_source)
        self.assertIn("create_payroll_running_balance_mutation", router_source)
        self.assertIn("Medewerker-inrichting aangepast", router_source)
        self.assertIn("Saldo-mutatie geboekt", router_source)
        self.assertIn("UPDATE payroll_employee_arrangements", records_source)
        self.assertIn("UPDATE payroll_running_balance_accounts", records_source)
        self.assertIn("INSERT INTO payroll_running_balance_mutations", records_source)
        self.assertIn("raw_status", records_source)
        self.assertIn("data-settings-edit-toggle", script)
        self.assertIn(".settings-inline-form", stylesheet)

    def test_timesheet_validation_replaces_existing_project_booking(self):
        correction_source = Path("apps/dashboard/timesheet_corrections.py").read_text(encoding="utf-8-sig")
        data_store = Path("apps/dashboard/data_store.py").read_text(encoding="utf-8-sig")
        migration = Path("migrations/047_deduplicate_project_time_bookings.sql").read_text(encoding="utf-8-sig")

        self.assertIn("DELETE FROM project_time_bookings", correction_source)
        self.assertIn("WHERE timesheet_inbox_id = %s", correction_source)
        self.assertIn("047_deduplicate_project_time_bookings.sql", data_store)
        self.assertIn("ROW_NUMBER() OVER", migration)
        self.assertIn("idx_project_time_bookings_unique_timesheet", migration)

    def test_timesheet_overview_shows_week_column(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn('data-sort-column="2">Week', template)
        self.assertIn("message.week_number_display", template)
        self.assertIn("week_number_sort", records_source)
        self.assertIn("parsed_week_number", records_source)

    def test_timesheet_validation_writes_to_matching_payroll_period_week(self):
        correction_source = Path("apps/dashboard/timesheet_corrections.py").read_text(encoding="utf-8-sig")

        self.assertIn("def _payroll_period_context", correction_source)
        self.assertIn("JOIN payroll_period_weeks w", correction_source)
        self.assertIn("payroll_period_id, payroll_period_week_id", correction_source)
        self.assertIn("INSERT INTO payroll_week_inputs", correction_source)
        self.assertIn("INSERT INTO payroll_week_input_days", correction_source)
        self.assertIn("INSERT INTO payroll_week_input_projects", correction_source)
        self.assertIn("payroll_cao_setting_id, payroll_period_id, work_date", correction_source)


class DashboardContextSafetyTests(unittest.TestCase):
    def test_dashboard_context_sections_fall_back_individually(self):
        def broken_loader():
            raise RuntimeError("kapot")

        self.assertEqual(dashboard_router._context_value("periods", "payroll_periods", [], broken_loader), [])

    def test_dashboard_context_logs_named_sections(self):
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")

        self.assertIn("DASHBOARD_CONTEXT_SECTION_ERROR", router_source)
        self.assertIn("selected_payroll_period", router_source)
        self.assertIn("timesheet_row", router_source)

    def test_timesheet_overview_links_directly_to_matching_payroll_period(self):
        shortcut = dashboard_router._timesheet_payroll_period_shortcut(
            [
                {
                    "workflow_stage": "loon",
                    "work_date": date(2026, 5, 18),
                    "received_at": datetime(2026, 5, 18, 8, 0),
                }
            ],
            [
                {
                    "id": 5,
                    "name": "Periode 5 2026",
                    "period_number": 5,
                    "raw_start_date": date(2026, 5, 4),
                    "raw_end_date": date(2026, 5, 31),
                }
            ],
        )

        self.assertEqual(shortcut["url"], "/dashboard/periods?period=5#periode-verloning")
        self.assertTrue(shortcut["matched"])
        self.assertIn("Periode 5 2026", shortcut["title"])

    def test_timesheet_period_context_still_matches_rows(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertNotIn("timesheet_payroll_period_shortcut.url", template)
        self.assertIn("Open loonperiode", template)
        self.assertIn("message.payroll_period_url", template)
        self.assertIn("_timesheet_payroll_period_shortcut(timesheet_stage_items, timesheet_payroll_periods)", router_source)
        self.assertIn('data_page in {"periods", "timesheets"}', router_source)
        self.assertIn('"raw_start_date"', records_source)

    def test_timesheet_rows_link_to_matching_payroll_periods(self):
        item = {"workflow_stage": "loon", "work_date": date(2026, 5, 18)}
        dashboard_router._attach_timesheet_payroll_period_link(
            item,
            [
                {
                    "id": 6,
                    "name": "Periode 06 25/05 - 21/06",
                    "period_number": 6,
                    "raw_start_date": date(2026, 5, 25),
                    "raw_end_date": date(2026, 6, 21),
                },
                {
                    "id": 5,
                    "name": "Periode 05 27/04 - 24/05",
                    "period_number": 5,
                    "raw_start_date": date(2026, 4, 27),
                    "raw_end_date": date(2026, 5, 24),
                },
            ],
        )

        self.assertEqual(item["payroll_period_url"], "/dashboard/periods?period=5#periode-verloning")
        self.assertEqual(item["payroll_period_title"], "Periode 05 27/04 - 24/05")

    def test_timesheet_overview_shows_row_period_link_for_payroll_and_archive(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")

        self.assertIn("Open loonperiode", template)
        self.assertIn("message.workflow_stage in ['loon', 'archief']", template)
        self.assertIn("message.payroll_period_url", template)
        self.assertIn("payroll_periods + archived_payroll_periods", router_source)
        self.assertIn('"archief": "Archief"', router_source)

    def test_timesheet_processed_status_moves_to_archive_stage(self):
        self.assertEqual(dashboard_router._timesheet_stage("processed"), "archief")
        self.assertEqual(dashboard_router._timesheet_stage("doorgestuurd_naar_loonadministratie"), "archief")
        self.assertEqual(dashboard_router._timesheet_stage("loon_te_berekenen"), "loon")

        tabs = dashboard_router._timesheet_workflow_tabs(
            [
                {"workflow_stage": "loon"},
                {"workflow_stage": "archief"},
            ],
            "loon",
        )

        self.assertNotIn("all", [tab["key"] for tab in tabs])
        self.assertEqual(next(tab for tab in tabs if tab["key"] == "loon")["count"], 1)
        self.assertEqual(next(tab for tab in tabs if tab["key"] == "archief")["count"], 1)

    def test_timesheet_workflow_starts_at_control_and_uses_validation_layer(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        script_source = Path("apps/dashboard/static/dashboard.js").read_text(encoding="utf-8-sig")
        stylesheet = Path("apps/dashboard/static/dashboard.css").read_text(encoding="utf-8-sig")

        self.assertNotIn("Alle taken", router_source)
        self.assertIn('workflow_stage: str = "controle"', router_source)
        self.assertIn('stage: str = "controle"', router_source)
        self.assertIn('{"controle", "valideren", "loon", "archief"} else "controle"', router_source)
        self.assertIn("Doorzetten naar loon berekenen", template)
        self.assertIn("selected_message.workflow_stage == 'valideren'", template)
        self.assertIn("Naar valideren", template)
        self.assertIn("data-can-send-to-payroll", template)
        self.assertIn("canSendToPayroll", script_source)
        self.assertIn("repeat(auto-fit, minmax(min(100%, 240px), 1fr))", stylesheet)
        self.assertIn("grid-auto-flow: row", stylesheet)

    def test_period_approval_finalizes_timesheets_and_archives_period(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertNotIn('/api/periods/{{ selected_payroll_period.id }}/approve', template)
        self.assertIn("Uitbetalen", template)
        self.assertIn("finalize_payroll_period_for_payment", router_source)
        self.assertIn("Loonperiode gevalideerd voor loonbetaling", router_source)
        self.assertIn("processed_timesheets", router_source)
        self.assertIn("def finalize_payroll_period_for_payment", records_source)
        self.assertIn("UPDATE whatsapp_timesheet_inbox", records_source)
        self.assertIn("UPDATE project_time_bookings", records_source)
        self.assertIn("UPDATE payroll_week_inputs", records_source)
        self.assertIn("SET status = 'Archief'", records_source)

    def test_payroll_period_controls_only_use_active_validation_rows(self):
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn("PAYROLL_VALIDATION_STATUSES", records_source)
        self.assertIn("def _active_period_payroll_status_condition", records_source)
        self.assertIn("get_payroll_week_result_summary", records_source)
        self.assertIn("list_payroll_period_exceptions", records_source)
        self.assertIn("get_payroll_week_input_summary", records_source)
        self.assertIn("LOWER(COALESCE(p.status, '')) <> 'archief'", records_source)
        self.assertIn("LOWER(REPLACE(COALESCE({alias}.status, ''), ' ', '_')) = ANY(%s)", records_source)

    def test_employee_arrangement_update_recalculates_only_open_payroll(self):
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn("recalculate_open_payroll_for_relation(row[0])", records_source)
        self.assertIn("def recalculate_open_payroll_for_relation", records_source)
        self.assertIn("LOWER(COALESCE(p.status, '')) <> 'archief'", records_source)
        self.assertIn("list(PAYROLL_VALIDATION_STATUSES)", records_source)
        self.assertIn("ON CONFLICT (payroll_week_input_id)", records_source)
        self.assertIn("'source', 'arrangement_update'", records_source)
        self.assertIn("DELETE FROM payroll_period_settlements", records_source)
        self.assertIn('period["period_settlements"] = list_payroll_period_settlements(period_id) if period.get("is_locked_for_payment") else []', records_source)

    def test_payroll_payment_approval_locks_timesheet_edits_until_period_restore(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        router_source = Path("apps/dashboard/router.py").read_text(encoding="utf-8-sig")
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")

        self.assertIn("get_timesheet_payroll_lock", router_source)
        self.assertIn("_timesheet_locked_response(timesheet_id", router_source)
        self.assertIn("status_code=423", router_source)
        self.assertIn("reopen_payroll_period_for_editing", router_source)
        self.assertIn("PAYROLL_LOCKED_STATUSES", records_source)
        self.assertIn("def get_timesheet_payroll_lock", records_source)
        self.assertIn("def is_payroll_period_locked_for_payment", records_source)
        self.assertIn("def reopen_payroll_period_for_editing", records_source)
        self.assertIn("payroll_sent_at = NULL", records_source)
        self.assertIn("Deze loonperiode is al gevalideerd voor loonbetaling en staat op slot", records_source)
        self.assertIn("selected_message.payroll_locked", template)
        self.assertIn("selected_payroll_period.is_locked_for_payment", template)
        self.assertIn("Gevalideerd voor loonbetaling", template)
        self.assertIn("disabled title", template)


class PayrollPeriodDetailSafetyTests(unittest.TestCase):
    def test_period_detail_uses_safe_defaults_and_warning(self):
        records_source = Path("apps/dashboard/records.py").read_text(encoding="utf-8-sig")
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")

        self.assertIn("def _empty_payroll_period_detail_defaults", records_source)
        self.assertIn("PAYROLL_PERIOD_DETAIL_WARNING", records_source)
        self.assertIn("detail_warning", records_source)
        self.assertIn("payroll-detail-warning", template)

    def test_period_controls_are_collapsible(self):
        template = Path("apps/dashboard/templates/dashboard.html").read_text(encoding="utf-8-sig")
        stylesheet = Path("apps/dashboard/static/dashboard.css").read_text(encoding="utf-8-sig")

        self.assertIn("payroll-control-details", template)
        self.assertIn("<summary>", template)
        self.assertIn("Controles en datamodel", template)
        self.assertIn(".payroll-control-details", stylesheet)


class PayrollFullYearTestSeedTests(unittest.TestCase):
    def test_full_year_test_seed_creates_missing_periods_and_weeks(self):
        migration = Path("migrations/039_full_year_test_payroll.sql").read_text(encoding="utf-8-sig")

        self.assertIn("generate_series(1, 13)", migration)
        self.assertIn("ON CONFLICT (year, period_number)", migration)
        self.assertIn("payroll_period_weeks", migration)
        self.assertIn("WHERE NOT EXISTS", migration)
        self.assertNotIn("whatsapp_timesheet_inbox", migration)
        self.assertNotIn("project_time_bookings", migration)

    def test_full_year_test_seed_runs_before_week_input_derivations(self):
        data_store = Path("apps/dashboard/data_store.py").read_text(encoding="utf-8-sig")

        self.assertNotIn("migrations/039_full_year_test_payroll.sql", data_store)
        self.assertNotIn("migrations/040_one_period_test_hours.sql", data_store)
        self.assertNotIn("migrations/041_dashboard_demo_payroll.sql", data_store)
        self.assertIn("migrations/033_payroll_week_inputs.sql", data_store)

    def test_one_period_test_hours_seed_populates_live_timesheets(self):
        migration = Path("migrations/040_one_period_test_hours.sql").read_text(encoding="utf-8-sig")

        self.assertIn("testdata_one_period", migration)
        self.assertIn("loon_te_berekenen", migration)
        self.assertIn("whatsapp_timesheet_inbox", migration)
        self.assertIn("project_time_bookings", migration)
        self.assertIn("test-one-period-2026", migration)
        self.assertIn("fallback_candidate", migration)

    def test_dashboard_demo_seed_populates_visible_relations_period_and_hours(self):
        migration = Path("migrations/041_dashboard_demo_payroll.sql").read_text(encoding="utf-8-sig")

        self.assertIn("dashboard_demo", migration)
        self.assertIn("dashboard-demo-candidate-001", migration)
        self.assertIn("dashboard-demo-principal-001", migration)
        self.assertIn("dashboard_demo_payroll", migration)
        self.assertIn("loon_te_berekenen", migration)
        self.assertIn("payroll_period_id", migration)
        self.assertIn("project_time_bookings", migration)
        self.assertIn("whatsapp_timesheet_inbox", migration)

if __name__ == "__main__":
    unittest.main()
