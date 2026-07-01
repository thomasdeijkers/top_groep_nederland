import re
from decimal import Decimal
from pathlib import Path

from apps.dashboard.payroll_calculations import PAYSLIP_SHEET_COLUMNS, PERIOD_SHEET_COLUMNS


TGN_TEMPLATE_PATH = Path("Hulp documenten/templates/TGN verloning template.xlsx")
TGN_TEMPLATE_WEEK_SHEETS = ("WK21", "WK22", "WK23", "WK24")
WEEK_SHEET_RE = re.compile(r"^WK\s*(\d{1,2})$", re.IGNORECASE)
PERIOD_SHEET = "periode"
PAYSLIP_SHEET = "loonstrook"
FOUNDATION_KEYWORDS = ("grondslag", "savg", "cao", "pensioen", "reservering")

WEEK_FIELD_ALIASES = {
    "employee_name": ("werknemer", "naam werknemer", "kandidaat"),
    "contract_hours": ("contract uren", "contracturen"),
    "worked_days": ("dagen gewerkt", "gewerkte dagen"),
    "worked_hours": ("uren gewerkt", "gewerkte uren"),
    "vacation_hours": ("vak opname", "vakantie", "vakantie-opname"),
    "sickness_hours": ("ziek opname", "ziekte", "ziekte-uren"),
    "rv_hours": ("rv opname", "roostervrij", "rv"),
    "kv_hours": ("kv/c doorbetaald", "kort verzuim", "kv"),
    "holiday_hours": ("fd doorbetaald", "feestdagen", "fd"),
    "net_amount": ("netto bedrag", "netto"),
    "commute_km": ("km woon-werk", "kilometers enkele reis", "enkele reis"),
    "work_km": ("km werk",),
    "total_km": ("totaal km", "kilometers"),
    "fuel_amount": ("brandstofbedrag", "brandstof"),
    "extra_reimbursement": ("extra declaratie", "vergoeding", "declaratie"),
    "net_advance": ("netto voorschot", "voorschot"),
    "remarks": ("opmerking", "opmerkingen"),
    "project_info": ("actueel project", "project"),
}

PERIOD_FIELD_ALIASES = {
    "employee_name": ("werknemer", "kandidaat"),
    "license_plate": ("kenteken",),
    "choice_budget": ("keuzebudget",),
    "phase": ("fase",),
    "pension_scheme": ("pensioen",),
    "contract_hours": ("contracturen", "contract uren"),
    "cao_name": ("cao",),
    "days_right": ("recht op dagen",),
    "configuration": ("inregeling",),
    "function_name": ("functie",),
    "gross_hourly_wage": ("bruto uurloon",),
    "gross_total": ("bruto totaal",),
    "reservations": ("reserveringen",),
}

PAYSLIP_FIELD_ALIASES = {
    "employee_name": ("werknemer",),
    "cao_name": ("cao",),
    "total_worked_days": ("totale dagen gewerkt", "dagen gewerkt"),
    "total_worked_hours": ("totale uren gewerkt", "uren gewerkt"),
    "total_vacation_hours": ("totaal vak", "vakantie"),
    "total_sickness_hours": ("totaal ziek", "ziek"),
    "total_rv_hours": ("totaal rv",),
    "total_kv_hours": ("totaal kv",),
    "total_holiday_hours": ("totaal fd",),
    "total_km": ("totaal kilometers", "totaal km"),
    "extra_reimbursements": ("extra declaraties", "extra vergoeding"),
    "already_received_net": ("reeds ontvangen netto",),
    "net_to_receive": ("nog te ontvangen netto", "nog te ontvangen netto loon"),
    "period_total": ("totaal 4 weken",),
    "wkr_reimbursements": ("wkr vergoedingen",),
    "payslip_advance": ("loonvoorschot strook",),
}


def analyze_payroll_workbook(path: str | Path) -> dict:
    workbook_path = Path(path)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is nodig om Excel-bestanden te analyseren.") from exc

    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    sheet_names = workbook.sheetnames
    week_tabs = []
    foundation_tabs = []
    mapped_fields = {}
    formulas = []
    warnings = []

    for sheet_name in sheet_names:
        normalized = _normalize(sheet_name)
        week_match = WEEK_SHEET_RE.match(sheet_name.strip())
        worksheet = workbook[sheet_name]
        if week_match:
            week_tabs.append({"sheet": sheet_name, "week_number": int(week_match.group(1))})
            mapped_fields[sheet_name] = _map_headers(worksheet, WEEK_FIELD_ALIASES)
        elif normalized == PERIOD_SHEET:
            mapped_fields[sheet_name] = _map_headers(worksheet, PERIOD_FIELD_ALIASES)
        elif normalized == PAYSLIP_SHEET:
            mapped_fields[sheet_name] = _map_headers(worksheet, PAYSLIP_FIELD_ALIASES)
        elif any(keyword in normalized for keyword in FOUNDATION_KEYWORDS):
            foundation_tabs.append(sheet_name)
            mapped_fields[sheet_name] = _map_headers(worksheet, PERIOD_FIELD_ALIASES | PAYSLIP_FIELD_ALIASES)
        formulas.extend(_collect_formulas(worksheet, sheet_name))

    if not week_tabs:
        warnings.append("Geen weektabs met patroon WKxx gevonden.")
    if not any(_normalize(name) == PERIOD_SHEET for name in sheet_names):
        warnings.append("Tabblad Periode niet gevonden.")
    if not any(_normalize(name) == PAYSLIP_SHEET for name in sheet_names):
        warnings.append("Tabblad Loonstrook niet gevonden.")

    return {
        "filename": workbook_path.name,
        "sheet_names": sheet_names,
        "week_tabs": sorted(week_tabs, key=lambda item: item["week_number"]),
        "period_sheet": _first_matching_sheet(sheet_names, PERIOD_SHEET),
        "payslip_sheet": _first_matching_sheet(sheet_names, PAYSLIP_SHEET),
        "foundation_sheets": foundation_tabs,
        "mapped_fields": mapped_fields,
        "formulas": formulas[:250],
        "formula_count": len(formulas),
        "warnings": warnings,
    }


def _first_matching_sheet(sheet_names: list[str], normalized_name: str) -> str | None:
    return next((name for name in sheet_names if _normalize(name) == normalized_name), None)


def _normalize(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _key(value: object) -> str:
    return _normalize(value)


def _map_headers(worksheet, aliases: dict[str, tuple[str, ...]]) -> dict:
    mapped = {}
    for row in worksheet.iter_rows(min_row=1, max_row=min(40, worksheet.max_row), max_col=min(80, worksheet.max_column)):
        for cell in row:
            text = _normalize(cell.value)
            if not text:
                continue
            for field_key, labels in aliases.items():
                if field_key in mapped:
                    continue
                if any(label in text for label in labels):
                    mapped[field_key] = {"label": str(cell.value), "cell": cell.coordinate}
    return mapped


def _collect_formulas(worksheet, sheet_name: str) -> list[dict]:
    formulas = []
    for row in worksheet.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                formulas.append({"sheet": sheet_name, "cell": cell.coordinate, "formula": value})
    return formulas


def build_payroll_output_workbook(path: str | Path, period: dict, use_tgn_template: bool = False) -> Path:
    if use_tgn_template and TGN_TEMPLATE_PATH.exists():
        return build_tgn_template_output_workbook(path, period, TGN_TEMPLATE_PATH)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is nodig om Excel-output te maken.") from exc

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    tabs = period.get("workbook_tabs") or [
        {"label": "Periode", "columns": PERIOD_SHEET_COLUMNS, "rows": period.get("period_sheet_rows", [])},
        {"label": "Loonstrook", "columns": PAYSLIP_SHEET_COLUMNS, "rows": period.get("payslip_sheet_rows", [])},
    ]
    tabs = [tab for tab in tabs if tab.get("columns")]
    if not tabs:
        tabs = [
            {
                "label": "Export",
                "columns": [{"label": "Melding", "key": "message"}],
                "rows": [{"message": "Geen regels klaar voor loonberekening."}],
            }
        ]
    sheet_map = {}
    used_titles = set()
    for tab in tabs:
        title = _safe_sheet_title(tab.get("label") or "Tabblad", used_titles)
        used_titles.add(title)
        sheet_map[title] = tab
        worksheet = workbook.create_sheet(title)
        _write_sheet(worksheet, tab.get("columns", []), tab.get("rows", []))
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        tab = sheet_map.get(worksheet.title, {})
        for index, column in enumerate(tab.get("columns", []), start=1):
            if column.get("hidden_in_excel"):
                worksheet.column_dimensions[get_column_letter(index)].hidden = True
        for column in worksheet.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(width + 2, 12), 34)
    workbook.save(output_path)
    return output_path


def build_tgn_template_output_workbook(path: str | Path, period: dict, template_path: str | Path = TGN_TEMPLATE_PATH) -> Path:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is nodig om Excel-output te maken.") from exc

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(template_path, data_only=False, read_only=False)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    tabs = period.get("workbook_tabs") or []
    week_tabs = [tab for tab in tabs if tab.get("kind") == "week"][:4]
    period_tab = next((tab for tab in tabs if tab.get("kind") == "period"), None)
    payslip_tab = next((tab for tab in tabs if tab.get("kind") == "payslip"), None)
    week_sheet_map = _rename_template_week_sheets(workbook, week_tabs)
    _replace_template_week_formula_refs(workbook, week_sheet_map)

    if period_tab:
        row_map = _fill_template_period_sheet(workbook["Periode"], period_tab.get("rows", []))
    else:
        row_map = {}
    for tab in week_tabs:
        sheet_name = week_sheet_map.get(tab.get("label")) or tab.get("label")
        if sheet_name in workbook.sheetnames:
            _fill_template_week_sheet(workbook[sheet_name], tab.get("rows", []), row_map)
    if payslip_tab and "Loonstrook" in workbook.sheetnames:
        _fill_template_payslip_notes(workbook["Loonstrook"], payslip_tab.get("rows", []), row_map)

    _reset_workbook_view(workbook)
    workbook.save(output_path)
    return output_path


def _rename_template_week_sheets(workbook, week_tabs: list[dict]) -> dict[str, str]:
    mapping = {}
    for old_name, tab in zip(TGN_TEMPLATE_WEEK_SHEETS, week_tabs):
        new_name = _safe_sheet_title(tab.get("label") or old_name, set(workbook.sheetnames) - {old_name})
        if old_name in workbook.sheetnames and old_name != new_name:
            workbook[old_name].title = new_name
        mapping[old_name] = new_name
        mapping[tab.get("label") or new_name] = new_name
    return mapping


def _replace_template_week_formula_refs(workbook, sheet_map: dict[str, str]) -> None:
    replacements = {
        old_name: new_name
        for old_name, new_name in sheet_map.items()
        if old_name in TGN_TEMPLATE_WEEK_SHEETS and old_name != new_name
    }
    if not replacements:
        return
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                for old_name, new_name in replacements.items():
                    value = value.replace(f"'{old_name}'!", f"'{new_name}'!")
                    value = value.replace(f"{old_name}!", f"'{new_name}'!")
                cell.value = value


def _template_employee_rows(worksheet) -> list[int]:
    rows = []
    for row_index in range(8, worksheet.max_row + 1):
        value = worksheet.cell(row_index, 2).value
        if value is None:
            continue
        if isinstance(value, str) and (value.startswith("=") or value.strip()):
            rows.append(row_index)
    return rows


def _clear_columns(worksheet, columns: tuple[str, ...], start_row: int = 8) -> None:
    try:
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        MergedCell = ()

    for row_index in range(start_row, worksheet.max_row + 1):
        for column in columns:
            cell = worksheet[f"{column}{row_index}"]
            if MergedCell and isinstance(cell, MergedCell):
                continue
            cell.value = None


def _clear_template_data_area(worksheet, start_row: int = 8) -> None:
    try:
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        MergedCell = ()

    for row in worksheet.iter_rows(min_row=start_row, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            if MergedCell and isinstance(cell, MergedCell):
                continue
            if _is_formula_cell(cell):
                continue
            cell.value = None


def _is_formula_cell(cell) -> bool:
    value = cell.value
    return cell.data_type == "f" or (isinstance(value, str) and value.startswith("="))


def _reset_workbook_view(workbook) -> None:
    try:
        from openpyxl.worksheet.views import Selection
    except ImportError:
        Selection = None

    workbook.active = 0
    for worksheet in workbook.worksheets:
        worksheet.sheet_view.topLeftCell = "A1"
        if Selection:
            worksheet.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
        else:
            for selection in worksheet.sheet_view.selection:
                selection.activeCell = "A1"
                selection.sqref = "A1"


def _fill_template_period_sheet(worksheet, rows: list[dict]) -> dict[str, int]:
    target_rows = _template_employee_rows(worksheet)
    _clear_template_period_inputs(worksheet)
    row_map = {}
    for target_row, source in zip(target_rows, rows):
        employee_name = source.get("employee_name") or ""
        row_map[_key(employee_name)] = target_row
        _set(worksheet, target_row, "B", employee_name)
        _set(worksheet, target_row, "C", source.get("license_plate"))
        _set(worksheet, target_row, "D", source.get("choice_budget"))
        _set(worksheet, target_row, "E", source.get("phase"))
        _set(worksheet, target_row, "F", source.get("pension_scheme"))
        _set(worksheet, target_row, "G", source.get("contract_hours"))
        _set(worksheet, target_row, "H", source.get("cao_name"))
        _set(worksheet, target_row, "I", source.get("days_right"))
        _set(worksheet, target_row, "J", source.get("configuration"))
        _set(worksheet, target_row, "K", source.get("function_name"))
        _set(worksheet, target_row, "L", source.get("gross_hourly_wage"))
        _set(worksheet, target_row, "M", source.get("gross_above_cao"))
        _set(worksheet, target_row, "T", source.get("reserve_vacation_days"))
        _set(worksheet, target_row, "V", source.get("reserve_adv"))
        _set(worksheet, target_row, "W", source.get("reserve_holiday"))
        _set(worksheet, target_row, "Y", source.get("tsf"))
        _set(worksheet, target_row, "AC", source.get("compensation_uta_days"))
        _set(worksheet, target_row, "AD", source.get("compensation_adv_days"))
        _set(worksheet, target_row, "AE", source.get("compensation_t"))
        _set(worksheet, target_row, "AI", source.get("rv_flex"))
        _set(worksheet, target_row, "BB", source.get("net_period_basis"))
        _set(worksheet, target_row, "BG", source.get("notes"))
    return row_map


def _clear_template_period_inputs(worksheet) -> None:
    _clear_template_data_area(worksheet)


def _fill_template_week_sheet(worksheet, rows: list[dict], row_map: dict[str, int]) -> None:
    _clear_template_week_inputs(worksheet)
    for source in rows:
        target_row = row_map.get(_key(source.get("employee_name")))
        if not target_row:
            continue
        _set(worksheet, target_row, "D", source.get("worked_days"))
        _set(worksheet, target_row, "E", source.get("worked_hours"))
        _set(worksheet, target_row, "F", source.get("vacation_hours"))
        _set(worksheet, target_row, "G", source.get("sickness_hours"))
        _set(worksheet, target_row, "H", source.get("rv_hours"))
        _set(worksheet, target_row, "I", source.get("kv_hours"))
        _set(worksheet, target_row, "J", source.get("holiday_hours"))
        _set(worksheet, target_row, "L", source.get("single_trip_km"))
        _set(worksheet, target_row, "M", source.get("work_km"))
        _set(worksheet, target_row, "P", source.get("extra_reimbursement"))
        _set(worksheet, target_row, "R", source.get("remarks"))
        _set(worksheet, target_row, "S", source.get("project_info"))


def _clear_template_week_inputs(worksheet) -> None:
    _clear_template_data_area(worksheet)


def _fill_template_payslip_notes(worksheet, rows: list[dict], row_map: dict[str, int]) -> None:
    _clear_template_payslip_notes(worksheet)
    for source in rows:
        target_row = row_map.get(_key(source.get("employee_name")))
        if not target_row:
            continue
        _set(worksheet, target_row, "P", source.get("notes"))


def _clear_template_payslip_notes(worksheet) -> None:
    _clear_template_data_area(worksheet)


def _set(worksheet, row: int, column: str, value) -> None:
    if value in (None, "", "-"):
        return
    worksheet[f"{column}{row}"].value = _excel_cell_value(value)


def _write_sheet(worksheet, columns: list[dict], rows: list[dict]) -> None:
    safe_columns = columns or [{"label": "Melding", "key": "message"}]
    worksheet.append([column.get("label", column.get("key", "")) for column in safe_columns])
    for row in rows:
        worksheet.append([_excel_cell_value(row.get(column.get("key", ""), "")) for column in safe_columns])


def _safe_sheet_title(value: str, used_titles: set[str]) -> str:
    base = re.sub(r"[:\\/?*\[\]]", "-", str(value or "Tabblad")).strip()[:31] or "Tabblad"
    title = base
    counter = 2
    while title in used_titles:
        suffix = f" {counter}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    return title


def _excel_cell_value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (dict, list, tuple, set)):
        return str(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text.startswith("="):
            return text
        parsed_number = _parse_excel_number(text)
        if parsed_number is not None:
            return parsed_number
    return value


def _parse_excel_number(value: str):
    text = value.strip()
    is_percent = text.endswith("%")
    text = text.replace(chr(8364), "").replace("â‚¬", "").replace("%", "").replace(" ", "")
    if not re.fullmatch(r"-?[0-9.,]+", text):
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    if is_percent:
        number = number / 100
    return int(number) if number.is_integer() else number
